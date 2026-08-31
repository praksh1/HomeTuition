"""Generate Sikshya's compact Nepal location/facility catalogue from the owner's workbook.

The workbook is source data only. This script intentionally drops `(unnamed)` facilities and
the `Unassigned` sheet; the product supplies one separate "Not specified" manual-entry choice.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def clean(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def unwanted(value: str) -> bool:
    return not value or value.casefold() in {"(unnamed)", "unnamed", "(unassigned)", "unassigned"}


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("usage: generate_nepal_education_data.py SOURCE.xlsx OUTPUT.json")

    source = Path(sys.argv[1]).resolve()
    output = Path(sys.argv[2]).resolve()
    book = load_workbook(source, read_only=True, data_only=True)
    province_names = [name for name in book.sheetnames if name not in {"Read Me", "Summary", "Unassigned"}]

    hierarchy: dict[str, dict[str, set[str]]] = {}
    facilities: list[list[object]] = []
    seen: set[tuple[str, str, str, str]] = set()

    for province_index, province in enumerate(province_names):
        sheet = book[province]
        rows = sheet.iter_rows(values_only=True)
        headers = [clean(value) for value in next(rows)]
        columns = {name: index for index, name in enumerate(headers)}
        required = {"District", "Local Level (Municipality)", "Facility Name"}
        if not required.issubset(columns):
            raise RuntimeError(f"{province}: missing columns {sorted(required - set(columns))}")

        province_tree = hierarchy.setdefault(province, {})
        for row in rows:
            district = clean(row[columns["District"]])
            local_level = clean(row[columns["Local Level (Municipality)"]])
            name = clean(row[columns["Facility Name"]])
            if unwanted(district) or unwanted(local_level) or unwanted(name):
                continue
            dedupe = (province.casefold(), district.casefold(), local_level.casefold(), name.casefold())
            if dedupe in seen:
                continue
            seen.add(dedupe)
            province_tree.setdefault(district, set()).add(local_level)
            nepali = clean(row[columns["Name (Nepali)"]]) if "Name (Nepali)" in columns else ""
            facility_type = clean(row[columns["Facility Type"]]) if "Facility Type" in columns else ""
            facilities.append([province_index, district, local_level, name, nepali, facility_type])

    payload = {
        "version": 1,
        "source": source.name,
        "provinces": [
            {
                "name": province,
                "districts": [
                    {"name": district, "localLevels": sorted(levels, key=str.casefold)}
                    for district, levels in sorted(hierarchy[province].items(), key=lambda item: item[0].casefold())
                ],
            }
            for province in province_names
        ],
        "facilities": facilities,
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"wrote {len(province_names)} provinces and {len(facilities)} named facilities to {output}")


if __name__ == "__main__":
    main()
